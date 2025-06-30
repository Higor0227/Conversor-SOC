import tkinter as tk
from tkinter import *
from tkinter import StringVar, filedialog
import PyPDF2
import PyPDF3
import openpyxl as xl
from tkinter import ttk
from tkinterdnd2 import DND_FILES, TkinterDnD
import statistics as stc
import os
from win32com import client
from PIL import Image
import win32com.client
from datetime import datetime


class EntryWithPlaceholder(tk.Entry):
    def __init__(self, master=None, placeholder="PLACEHOLDER", color='grey'):
        super().__init__(master)

        self.placeholder = placeholder
        self.placeholder_color = color
        self.default_fg_color = self['fg']

        self.bind("<FocusIn>", self.foc_in)
        self.bind("<FocusOut>", self.foc_out)

        self.put_placeholder()

    def put_placeholder(self):
        self.insert(0, self.placeholder)
        self['fg'] = self.placeholder_color

    def foc_in(self, *args):
        if self['fg'] == self.placeholder_color:
            self.delete('0', 'end')
            self['fg'] = self.default_fg_color

    def foc_out(self, *args):
        if not self.get():
            self.put_placeholder()


png = [0, 0, 0]


def pdffinal(current_path):
    global png
    # LER NxM DA PLANILHA template
    numeros = []
    for valor_entry in [valor1, valor2, valor3]:
        valor_text = valor_entry.get().replace(',', '.')
        try:
            valor_float = float(valor_text)
            numeros.append(valor_float)
        except ValueError:
            pass

    if len(numeros) == 1:
        media = numeros[0]
        mediana = stc.median(numeros)
        menor = numeros[0]
        desvpad = numeros[0]
        liminf = numeros[0]
        limsup = numeros[0]
        coefvaria = 0
    else:
        media = stc.mean(numeros)
        mediana = stc.median(numeros)
        menor = min(numeros)
        desvpad = stc.stdev(numeros)
        coefvaria = desvpad / media
        liminf = media - desvpad
        limsup = media + desvpad

    if coefvaria < 0.25:
        wb = xl.load_workbook(filename='template.xlsx')
        ws = wb.active
    else:
        wb = xl.load_workbook(filename='template hetero.xlsx')
        ws = wb.active

    if len(numeros) == 3:
        ws['A2'] = item.get()
        ws['B2'] = insumo.get()
        ws['A7'] = 'Valor (' + unidade.get() + ')'
        ws['B3'] = fornecedor1.get()
        ws['C3'] = fornecedor2.get()
        ws['D3'] = fornecedor3.get()
        ws['B4'] = data1.get()
        ws['C4'] = data2.get()
        ws['D4'] = data3.get()
        fonte = f"{fornecedor1.get()}: {fonte1.get()}\n{fornecedor2.get()}: {fonte2.get()}\n{fornecedor3.get()}: {fonte3.get()}"

        if frete1.get() == "Valor":
            ws['B5'] = 'Valor do frete: R$' + valor1_entry.get()
        else:
            ws['B5'] = frete1.get()
        if frete2.get() == "Valor":
            ws['C5'] = 'Valor do frete: R$' + valor2_entry.get()
        else:
            ws['C5'] = frete2.get()
        if frete3.get() == "Valor":
            ws['D5'] = 'Valor do frete: R$' + valor3_entry.get()
        else:
            ws['D5'] = frete3.get()

        if mo1.get() == "Sim":
            ws['B6'] = 'Valor da mão de obra: R$' + mo1_entry.get()
        if mo1.get() == "Mão de obra inclusa":
            ws['B6'] = "Mão de obra inclusa"
        else:
            ws['B6'] = '-'
        if mo2.get() == "Sim":
            ws['C6'] = 'Valor da mão de obra: R$' + mo2_entry.get()
        if mo2.get() == "Mão de obra inclusa":
            ws['C6'] = "Mão de obra inclusa"
        else:
            ws['C6'] = '-'
        if mo3.get() == "Sim":
            ws['D6'] = 'Valor da mão de obra: R$' + mo3_entry.get()
        if mo3.get() == "Mão de obra inclusa":
            ws['D6'] = "Mão de obra inclusa"
        else:
            ws['D6'] = '-'

    elif len(numeros) == 2:
        ws['A2'] = item.get()
        ws['B2'] = insumo.get()
        ws['A7'] = 'Valor (' + unidade.get() + ')'
        ws['B3'] = fornecedor1.get()
        ws['C3'] = fornecedor2.get()
        ws['B4'] = data1.get()
        ws['C4'] = data2.get()
        fonte = f"{fornecedor1.get()}: {fonte1.get()}\n{fornecedor2.get()}: {fonte2.get()}"

        if frete1.get() == "Valor":
            ws['B5'] = 'Valor do frete: R$' + valor1_entry.get()
        else:
            ws['B5'] = frete1.get()
        if frete2.get() == "Valor":
            ws['C5'] = 'Valor do frete: R$' + valor2_entry.get()
        else:
            ws['C5'] = frete2.get()

        if mo1.get() == "Sim":
            ws['B6'] = 'Valor da mão de obra: R$' + mo1_entry.get()
        if mo1.get() == "Mão de obra inclusa":
            ws['B6'] = "Mão de obra inclusa"
        else:
            ws['B6'] = '-'
        if mo2.get() == "Sim":
            ws['C6'] = 'Valor da mão de obra: R$' + mo2_entry.get()
        if mo2.get() == "Mão de obra inclusa":
            ws['C6'] = "Mão de obra inclusa"
        else:
            ws['C6'] = '-'

    else:
        ws['A2'] = item.get()
        ws['B2'] = insumo.get()
        ws['A7'] = 'Valor (' + unidade.get() + ')'
        ws['B3'] = fornecedor1.get()
        ws['B4'] = data1.get()
        fonte = f"{fornecedor1.get()}: {fonte1.get()}"

        if frete1.get() == "Valor":
            ws['B5'] = 'Valor do frete: R$' + valor1_entry.get()
        else:
            ws['B5'] = frete1.get()

        if mo1.get() == "Sim":
            ws['B6'] = 'Valor da mão de obra: R$' + mo1_entry.get()
        if mo1.get() == "Mão de obra inclusa":
            ws['B6'] = "Mão de obra inclusa"
        else:
            ws['B6'] = '-'

    if coefvaria < 0.25:
        ws['B15'] = 'Homogêneo'
    else:
        ws['B15'] = 'Heterogêneo'

    if v1.get() == 1:
        check = True
    else:
        check = False

    if check:
        ws['B8'] = numeros[0]
    else:
        if coefvaria < 0.25:
            ws['B8'] = media
        else:
            ws['B8'] = mediana

    if len(numeros) == 3:
        ws['B7'] = numeros[0]
        ws['C7'] = numeros[1]
        ws['D7'] = numeros[2]
    elif len(numeros) == 2:
        ws['B7'] = numeros[0]
        ws['C7'] = numeros[1]
    else:
        ws['B7'] = numeros[0]

    ws['B10'] = media
    ws['B11'] = mediana
    ws['B12'] = menor
    ws['B13'] = desvpad
    ws['B14'] = coefvaria
    ws['B16'] = liminf
    ws['B17'] = limsup

    wb.save('documento.xlsx')
    valor_final = ws['B8'].value
    valor_final = round(valor_final, 2)
    valor_final = str(valor_final)
    valor_final = valor_final.replace(".", ",")
    wb = client.Dispatch("Excel.Application")

    # Abrir o Excel e converter para PDF
    excel = client.Dispatch("Excel.Application")
    sheets = excel.Workbooks.Open(os.path.abspath('documento.xlsx'))
    work_sheets = sheets.Worksheets[0]
    work_sheets.ExportAsFixedFormat(0, current_path + 'saida.pdf')
    sheets.Close(False)

    # Abrir o PDF de saída
    pdf_file = open(current_path + 'saida.pdf', 'rb')
    pdf_reader = PyPDF3.PdfFileReader(pdf_file)
    pdf_writer = PyPDF3.PdfFileWriter()

    # Copiar todas as páginas, exceto a segunda, para o objeto PDFWriter
    for page_num in range(pdf_reader.getNumPages()):
        if page_num != 1:  # Não inclua a segunda página (índice 1)
            page = pdf_reader.getPage(page_num)
            pdf_writer.addPage(page)

    # Salvar o novo arquivo PDF sem a segunda página
    with open(current_path + 'saida_sem_segunda.pdf', 'wb') as output_pdf:
        pdf_writer.write(output_pdf)

    documentofinal = PyPDF2.PdfMerger()

    documentofinal.append(PyPDF2.PdfReader(current_path + 'saida_sem_segunda.pdf', 'rb'))

    if png[0]:
        documentofinal.append(PyPDF2.PdfReader(current_path + "image1.pdf", 'rb'))
    else:
        documentofinal.append(PyPDF2.PdfReader(arquivo_entry1.get().translate({ord("{"): None, ord("}"): None})))

    if len(numeros) >= 2:
        if png[1]:
            documentofinal.append(PyPDF2.PdfReader(current_path + "image2.pdf", 'rb'))
        else:
            documentofinal.append(PyPDF2.PdfReader(arquivo_entry2.get().translate({ord("{"): None, ord("}"): None})))

    if len(numeros) == 3:
        if png[2]:
            documentofinal.append(PyPDF2.PdfReader(current_path + "image3.pdf", 'rb'))
        else:
            documentofinal.append(PyPDF2.PdfReader(arquivo_entry3.get().translate({ord("{"): None, ord("}"): None})))

    os.remove(current_path + 'documento.xlsx')
    pdf_file.close()
    os.remove(current_path + 'saida.pdf')
    os.remove(current_path + 'saida_sem_segunda.pdf')
    if png[0]:
        os.remove(current_path + 'image1.pdf')

    if len(numeros) >= 2:
        if png[1]:
            os.remove(current_path + 'image2.pdf')

    if len(numeros) == 3:
        if png[2]:
            os.remove(current_path + 'image3.pdf')

    pdf_file.close()

    png = [0, 0, 0]

    diretorio = entrada_diretorio.get()
    if diretorio[-1] != '/':
        diretorio = entrada_diretorio.get() + '/'

    documentofinal.write(f"{diretorio}in.pdf")

    with open(diretorio + 'sga.txt', 'w') as f:
        f.write(valor_final)
        f.write(';')
        f.write(datas())
        f.write(';')
        f.write(fonte)
        f.write(';')
        f.write(diretorio.rstrip("/") + "\\in.pdf")


def clear_entries():
    insumo.delete(0, 'end')
    insumo.put_placeholder()
    fornecedor1.delete(0, 'end')
    fornecedor1.put_placeholder()
    data1.delete(0, 'end')
    data1.put_placeholder()
    valor1.delete(0, 'end')
    valor1.put_placeholder()
    frete1.set("Selecione o frete")
    valor1_entry.delete(0, 'end')
    valor1_entry.put_placeholder()
    mo1.set("Mão de obra?")
    mo1_entry.delete(0, 'end')
    mo1_entry.put_placeholder()
    arquivo_entry1.delete(0, 'end')

    item.delete(0, 'end')
    item.put_placeholder()
    fornecedor2.delete(0, 'end')
    fornecedor2.put_placeholder()
    data2.delete(0, 'end')
    data2.put_placeholder()
    valor2.delete(0, 'end')
    valor2.put_placeholder()
    frete2.set("Selecione o frete")
    valor2_entry.delete(0, 'end')
    valor2_entry.put_placeholder()
    mo2.set("Mão de obra?")
    mo2_entry.delete(0, 'end')
    mo2_entry.put_placeholder()
    arquivo_entry2.delete(0, 'end')

    unidade.delete(0, 'end')
    unidade.put_placeholder()
    fornecedor3.delete(0, 'end')
    fornecedor3.put_placeholder()
    data3.delete(0, 'end')
    data3.put_placeholder()
    valor3.delete(0, 'end')
    valor3.put_placeholder()
    frete3.set("Selecione o frete")
    valor3_entry.delete(0, 'end')
    valor3_entry.put_placeholder()
    mo3.set("Mão de obra?")
    mo3_entry.delete(0, 'end')
    mo3_entry.put_placeholder()
    arquivo_entry3.delete(0, 'end')

    entrada_diretorio.delete(0, 'end')

    v1.set(0)


def desbug():
    current_path = os.getcwd()  # Obtém o diretório de trabalho atual

    try:
        # Fecha o Excel em segundo plano
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Quit()
    except Exception as e:
        pass

    files_to_remove = ['documento.xlsx', 'saida.pdf', 'saida_sem_segunda.pdf', 'image1.pdf', 'image2.pdf', 'image3.pdf']

    for file_name in files_to_remove:
        file_path = os.path.join(current_path, file_name)
        if os.path.exists(file_path):
            os.remove(file_path)
        else:
            pass


def on_frete1_select(event):
    selected_option = frete1.get()

    if selected_option == "Valor":
        valor1_entry.pack()  # Exibe a caixa de texto
    else:
        valor1_entry.pack_forget()  # Oculta a caixa de texto


def on_frete2_select(event):
    selected_option = frete2.get()

    if selected_option == "Valor":
        valor2_entry.pack()  # Exibe a caixa de texto
    else:
        valor2_entry.pack_forget()  # Oculta a caixa de texto


def on_frete3_select(event):
    selected_option = frete3.get()

    if selected_option == "Valor":
        valor3_entry.pack()  # Exibe a caixa de texto
    else:
        valor3_entry.pack_forget()  # Oculta a caixa de texto


def on_mo1_select(event):
    selected_option = mo1.get()

    if selected_option == "Sim":
        mo1_entry.pack()  # Exibe a caixa de texto
    else:
        mo1_entry.pack_forget()  # Oculta a caixa de texto


def on_mo2_select(event):
    selected_option = mo2.get()

    if selected_option == "Sim":
        mo2_entry.pack()  # Exibe a caixa de texto
    else:
        mo2_entry.pack_forget()  # Oculta a caixa de texto


def on_mo3_select(event):
    selected_option = mo3.get()

    if selected_option == "Sim":
        mo3_entry.pack()  # Exibe a caixa de texto
    else:
        mo3_entry.pack_forget()  # Oculta a caixa de texto


# Crie a janela principal
janela = TkinterDnD.Tk()
janela.title("Conversor")
janela['bg'] = '#0D214F'
janela.geometry('980x600')
janela.resizable(False, False)

s = ttk.Style()
s.configure('My.TFrame', background='#0D214F')

png = [0, 0, 0]

# Crie três frames para as três colunas
frame1 = ttk.Frame(janela, style='My.TFrame')
frame2 = ttk.Frame(janela, style='My.TFrame')
frame3 = ttk.Frame(janela, style='My.TFrame')
frame4 = ttk.Frame(janela, style='My.TFrame')
frame5 = ttk.Frame(janela, style='My.TFrame')
frame6 = ttk.Frame(janela, style='My.TFrame')

# Cria as opções do frete
opcoes_frete = ["Selecione o frete", "Frete CIF", "Frete FOB", "Frete Indisponível", "Frete incluso", "Valor"]
opcoes_mo = ["Mão de obra?", "Sim", "Não", "Mão de obra inclusa"]


def drop1(event):
    var1.set(event.data)
    if arquivo_entry1.get().endswith((".jpg", ".jpeg", ".png")):
        convert_image_to_pdf(event.data, "image1.pdf")
        png[0] = 1


def drop2(event):
    var2.set(event.data)
    if event.data.endswith((".jpg", ".jpeg", ".png")):
        convert_image_to_pdf(event.data, "image2.pdf")
        png[1] = 1


def drop3(event):
    var3.set(event.data)
    if event.data.endswith((".jpg", ".jpeg", ".png")):
        convert_image_to_pdf(event.data, "image3.pdf")
        png[2] = 1


def convert_image_to_pdf(image_path, output_pdf):
    image = Image.open(image_path)
    image.save(output_pdf, "PDF")


# Cria a primeira coluna
insumo = EntryWithPlaceholder(frame1, 'Insumo')
insumo.pack(pady=25)
fornecedor1 = EntryWithPlaceholder(frame1, 'Fornecedor')
fornecedor1.pack()
data1 = EntryWithPlaceholder(frame1, 'Data')
data1.pack()
valor1 = EntryWithPlaceholder(frame1, 'Valor do insumo')
valor1.pack()
frete1 = ttk.Combobox(frame1, values=opcoes_frete)
frete1.current(0)
frete1.pack()
valor1_entry = EntryWithPlaceholder(frame1, 'Valor do frete')
frete1.bind("<<ComboboxSelected>>", on_frete1_select)
mo1 = ttk.Combobox(frame1, values=opcoes_mo)
mo1.current(0)
mo1.pack()
mo1_entry = EntryWithPlaceholder(frame1, 'Valor da MO')
mo1.bind("<<ComboboxSelected>>", on_mo1_select)
var1 = StringVar()
arquivo_entry1 = Entry(frame1, textvar=var1, width=40)
arquivo_entry1.pack(side='bottom')
arquivo_entry1.drop_target_register(DND_FILES)
arquivo_entry1.dnd_bind('<<Drop>>', drop1)
label1 = tk.Label(frame1, text="Arraste e solte arquivos aqui:", bg='#0D214F', fg='white')
label1.pack(pady=15, side='bottom')
fonte1 = EntryWithPlaceholder(frame1, 'Fonte')
fonte1.pack(pady=15, side='bottom')
fonte1.config(width=39)

# Cria a segunda coluna
item = EntryWithPlaceholder(frame2, 'Número Item')
item.pack(pady=25)
fornecedor2 = EntryWithPlaceholder(frame2, 'Fornecedor')
fornecedor2.pack()
data2 = EntryWithPlaceholder(frame2, 'Data')
data2.pack()
valor2 = EntryWithPlaceholder(frame2, 'Valor do insumo')
valor2.pack()
frete2 = ttk.Combobox(frame2, values=opcoes_frete)
frete2.current(0)
frete2.pack()
valor2_entry = EntryWithPlaceholder(frame2, 'Valor do frete')
frete2.bind("<<ComboboxSelected>>", on_frete2_select)
mo2 = ttk.Combobox(frame2, values=opcoes_mo)
mo2.current(0)
mo2.pack()
mo2_entry = EntryWithPlaceholder(frame2, 'Valor da MO')
mo2.bind("<<ComboboxSelected>>", on_mo2_select)
var2 = StringVar()
arquivo_entry2 = Entry(frame2, textvar=var2, width=40)
arquivo_entry2.pack(side='bottom')
arquivo_entry2.drop_target_register(DND_FILES)
arquivo_entry2.dnd_bind('<<Drop>>', drop2)
label2 = tk.Label(frame2, text="Arraste e solte arquivos aqui:", bg='#0D214F', fg='white')
label2.pack(pady=15, side='bottom')
fonte2 = EntryWithPlaceholder(frame2, 'Fonte')
fonte2.pack(pady=15, side='bottom')
fonte2.config(width=39)

# Cria a terceira coluna
unidade = EntryWithPlaceholder(frame3, 'Unidade')
unidade.pack(pady=25)
fornecedor3 = EntryWithPlaceholder(frame3, 'Fornecedor')
fornecedor3.pack()
data3 = EntryWithPlaceholder(frame3, 'Data')
data3.pack()
valor3 = EntryWithPlaceholder(frame3, 'Valor do insumo')
valor3.pack()
frete3 = ttk.Combobox(frame3, values=opcoes_frete)
frete3.current(0)
frete3.pack()
valor3_entry = EntryWithPlaceholder(frame3, 'Valor do frete')
frete3.bind("<<ComboboxSelected>>", on_frete3_select)
mo3 = ttk.Combobox(frame3, values=opcoes_mo)
mo3.current(0)
mo3.pack()
mo3_entry = EntryWithPlaceholder(frame3, 'Valor da MO')
mo3.bind("<<ComboboxSelected>>", on_mo3_select)
var3 = StringVar()
arquivo_entry3 = Entry(frame3, textvar=var3, width=40)
arquivo_entry3.pack(side='bottom')
arquivo_entry3.drop_target_register(DND_FILES)
arquivo_entry3.dnd_bind('<<Drop>>', drop3)
label3 = tk.Label(frame3, text="Arraste e solte arquivos aqui:", bg='#0D214F', fg='white')
label3.pack(pady=15, side='bottom')
fonte3 = EntryWithPlaceholder(frame3, 'Fonte')
fonte3.pack(pady=15, side='bottom')
fonte3.config(width=39)

label_diretorio = tk.Label(frame4, text="Digite o diretório onde deseja criar o arquivo:", background='#0D214F',
                           foreground='white')
label_diretorio.pack(pady=0)
entrada_diretorio = tk.Entry(frame4, width=70)
entrada_diretorio.pack()
current_path = os.getcwd()


def datas():
    # Definir lista de datas como strings
    datas_str = [data1.get(), data2.get(), data3.get()]

    # Filtrar e converter strings válidas para objetos datetime
    datas = []
    for data_str in datas_str:
        if data_str != "Data":
            try:
                datas.append(datetime.strptime(data_str, "%d/%m/%Y"))
            except ValueError:
                continue

    # Verificar se há datas válidas e encontrar a menor data
    if datas:
        prim_data = min(datas).strftime('%d/%m/%Y')
    else:
        prim_data = "Nenhuma data válida"

    return prim_data


def ajuda():
    janela_2 = tk.Toplevel()
    janela_2.title("Janela de Ajuda")
    janela_2.geometry("420x650")
    janela_2.resizable(False, False)

    # Texto da apresentação do programa
    apresentacao_label = tk.Label(janela_2, text="Apresentação do Programa", font=("Arial", 10, "bold"))
    apresentacao_label.pack(anchor='w', pady=(10, 0), padx=10)
    apresentacao_text = tk.Label(janela_2,
                                 text="    Olá,\n    Sou o Daniel, ex-estagiário da DEA (seção de orçamento e custos), e em colaboração com a equipe da SOC, desenvolvi este programa para automatizar um processo que antes era realizado de forma menos eficiente.\n    Embora o programa ainda esteja em estágio inicial, há muitas oportunidades para melhorias e novos recursos. Se você tiver sugestões ou quiser contribuir, o arquivo .py do Conversor está disponível na minha pasta de colaboradores (Daniel dos Anjos).\n    Fique à vontade para entrar em contato pelo WhatsApp para discutirmos mais sobre o programa! Meu número é 48 99105-1504.",
                                 font=("Arial", 10), justify="left", wraplength=400)
    apresentacao_text.pack(anchor='w', pady=(0, 10), padx=10)

    # Texto sobre o programa
    sobre_label = tk.Label(janela_2, text="Sobre o Programa", font=("Arial", 10, "bold"))
    sobre_label.pack(anchor='w', pady=(10, 0), padx=10)
    sobre_text = tk.Label(janela_2,
                          text="    O objetivo do programa é criar a planilha in e juntá-la com os pdfs dos respectivos orçamentos, utilizada para cadastro no SGA e Volare.\n    O Conversor está atualmente na versão 7, com alguns bugs na hora de formar o pdf, devido à falta de permissões do programa para alterar o disco C:.",
                          font=("Arial", 10), justify="left", wraplength=400)
    sobre_text.pack(anchor='w', pady=(0, 10), padx=10)

    # Texto de como usar o programa
    como_usar_label = tk.Label(janela_2, text="Como Usar", font=("Arial", 10, "bold"))
    como_usar_label.pack(anchor='w', pady=(10, 0), padx=10)
    como_usar_text = tk.Label(janela_2,
                              text="    - Preencha todos os campos com suas devidas informações;\n    - Arraste e solte os arquivos para os campos indicados;\n    - O botão anti-pânico apaga os arquivos temporários do Conversor;\n    - Clique no botão Criar documento in para finalizar o procedimento;\n    - O Botão Limpar limpa as entradas já preenchidas.",
                              font=("Arial", 10), justify="left", wraplength=400)
    como_usar_text.pack(anchor='w', pady=(0, 10), padx=10)

    # Texto de dicas
    como_usar_label = tk.Label(janela_2, text="Dicas", font=("Arial", 10, "bold"))
    como_usar_label.pack(anchor='w', pady=(10, 0), padx=10)
    como_usar_text = tk.Label(janela_2,
                              text="    Dependendo da situação, o programa pode bugar e não funcionar mais. Nessas situações, feche o programa e procure a pasta de instalações do Conversor. Na pasta, apague os arquivos temporários: documento, saida e saida_sem_segunda.\n    Reinicie o programa normalmente.",
                              font=("Arial", 10), justify="left", wraplength=400)
    como_usar_text.pack(anchor='w', pady=(0, 10), padx=10)


botao_limpar = tk.Button(frame6, text="Limpar", command=clear_entries)
botao_limpar.pack(pady=10)  # Add some spacing below the button

botao_ajuda = tk.Button(frame6, text="Ajuda", command=ajuda)
botao_ajuda.pack(pady=10)  # Add some spacing below the button

botao_desbug = tk.Button(frame6, text="Botão anti-pânico", command=desbug)
botao_desbug.pack(pady=10)  # Add some spacing below the button

botao_criar = tk.Button(frame4, text="Criar documento in", command=lambda: pdffinal(current_path + "\\"))
botao_criar.pack(pady=10)  # Adicionei um pequeno espaço abaixo do botão

# Check do insumo especial de marca
v1 = tk.IntVar()
check = tk.Checkbutton(frame5, text='insumo de marca?', onvalue=1, offvalue=0, var=v1)
check.pack(pady=10)

# Texto acima do botão
texto_autor = tk.Label(frame4, text="Feito por Daniel dos Anjos e Joseph Petrassem", bg='#0D214F', fg='white')
texto_autor.pack(side='bottom')

# Organize os frames em uma grade
frame1.grid(row=0, column=0, padx=10, pady=10)
frame2.grid(row=0, column=1, padx=10, pady=10)
frame3.grid(row=0, column=2, padx=10, pady=10)
frame5.grid(row=1, column=0, padx=10, pady=10)
frame4.grid(row=2, column=1, padx=10, pady=10)
frame6.grid(row=2, column=2, padx=10, pady=10)

# Execute o loop principal da interface gráfica
janela.mainloop()

# =============================================================================================
# para duvidas, enviar pix para 09957068911 e depois mandar msg para 48 991051504!
# =============================================================================================